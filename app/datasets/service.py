"""Dataset service — upload, schema extraction, and CRUD orchestration."""

import json
import re
import secrets
from datetime import datetime
from pathlib import Path

import structlog
from fastapi import UploadFile

from app.config import Settings
from app.core.exceptions import (
    DatasetNotFoundError,
    FileTooLargeError,
    SchemaExtractionError,
    UnsupportedFileTypeError,
)
from app.datasets.repository import DatasetRepository
from app.datasets.schemas import DatasetResponse, SchemaColumn
from app.db.duckdb import DuckDBManager
from app.db.metadata import DatasetRecord

logger = structlog.get_logger()

ALLOWED_EXTENSIONS = {".csv", ".tsv", ".xlsx", ".xls"}


def _sanitize_table_name(name: str) -> str:
    """Convert a filename into a safe DuckDB table identifier."""
    stem = Path(name).stem
    clean = re.sub(r"[^a-zA-Z0-9_]", "_", stem).lower().strip("_")
    if clean and clean[0].isdigit():
        clean = f"t_{clean}"
    return clean or "dataset"


def _generate_dataset_id() -> str:
    """Generate a short unique dataset ID."""
    return f"ds_{secrets.token_hex(4)}"


class DatasetService:
    """Orchestrates dataset upload, schema extraction, and lifecycle."""

    def __init__(
        self,
        db: DuckDBManager,
        repository: DatasetRepository,
        settings: Settings,
    ) -> None:
        self._db = db
        self._repo = repository
        self._settings = settings

    async def upload(
        self,
        file: UploadFile,
        name: str | None = None,
        description: str | None = None,
    ) -> DatasetResponse:
        """Upload a file, extract schema, and persist metadata."""
        filename = file.filename or "unknown"
        ext = Path(filename).suffix.lower()

        if ext not in ALLOWED_EXTENSIONS:
            raise UnsupportedFileTypeError(filename)

        # Read file content and check size
        content = await file.read()
        size_mb = len(content) / (1024 * 1024)
        if size_mb > self._settings.MAX_UPLOAD_SIZE_MB:
            raise FileTooLargeError(size_mb, self._settings.MAX_UPLOAD_SIZE_MB)

        dataset_id = _generate_dataset_id()
        table_name = _sanitize_table_name(name or filename)

        # For Excel files, convert to CSV via openpyxl
        if ext in {".xlsx", ".xls"}:
            csv_path = await self._convert_excel_to_csv(content, dataset_id)
        else:
            csv_path = self._save_file(content, dataset_id, ext)

        # Extract schema via DuckDB
        try:
            schema_rows = await self._db.describe_table(csv_path)
            sample_rows = await self._db.sample_rows(csv_path, limit=5)
            row_count = await self._db.count_rows(csv_path)
        except Exception as e:
            logger.error("schema_extraction_failed", file=csv_path, error=str(e))
            raise SchemaExtractionError(csv_path, str(e)) from e

        # Build schema info
        schema_columns = []
        for col in schema_rows:
            col_name = col["column_name"]
            col_type = col["column_type"]
            samples = [row.get(col_name) for row in sample_rows if row.get(col_name) is not None]
            schema_columns.append(SchemaColumn(column=col_name, type=col_type, sample=samples[:5]))

        record = DatasetRecord(
            id=dataset_id,
            name=name or Path(filename).stem,
            description=description,
            table_name=table_name,
            file_path=csv_path,
            schema_json=json.dumps([c.model_dump() for c in schema_columns], default=str),
            sample_json=json.dumps(sample_rows, default=str),
            row_count=row_count,
            file_size_bytes=len(content),
            created_at=datetime.now(),
        )
        await self._repo.save(record)

        logger.info("dataset_uploaded", dataset_id=dataset_id, rows=row_count)

        return DatasetResponse(
            id=record.id,
            name=record.name,
            description=record.description,
            table_name=record.table_name,
            row_count=record.row_count,
            schema_info=schema_columns,
            file_size_bytes=record.file_size_bytes,
            created_at=record.created_at,
        )

    async def list_datasets(self) -> list[DatasetResponse]:
        """List all datasets."""
        records = await self._repo.list_all()
        return [self._record_to_response(r) for r in records]

    async def get_dataset(self, dataset_id: str) -> DatasetResponse:
        """Get a single dataset by ID."""
        record = await self._repo.get_by_id(dataset_id)
        if not record:
            raise DatasetNotFoundError(dataset_id)
        return self._record_to_response(record)

    async def delete_dataset(self, dataset_id: str) -> None:
        """Delete a dataset's file and metadata."""
        record = await self._repo.get_by_id(dataset_id)
        if not record:
            raise DatasetNotFoundError(dataset_id)

        # Remove the file
        file_path = Path(record.file_path)
        if file_path.exists():
            file_path.unlink()

        await self._repo.delete_by_id(dataset_id)
        logger.info("dataset_deleted", dataset_id=dataset_id)

    def _save_file(self, content: bytes, dataset_id: str, ext: str) -> str:
        """Save raw file content to the upload directory."""
        upload_dir = Path(self._settings.UPLOAD_DIR)
        upload_dir.mkdir(parents=True, exist_ok=True)
        file_path = upload_dir / f"{dataset_id}{ext}"
        file_path.write_bytes(content)
        return str(file_path)

    async def _convert_excel_to_csv(self, content: bytes, dataset_id: str) -> str:
        """Convert Excel bytes to CSV and save to upload directory."""
        import csv
        import io

        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active

        upload_dir = Path(self._settings.UPLOAD_DIR)
        upload_dir.mkdir(parents=True, exist_ok=True)
        csv_path = upload_dir / f"{dataset_id}.csv"

        with open(csv_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            for row in ws.iter_rows(values_only=True):
                writer.writerow(row)

        wb.close()
        return str(csv_path)

    @staticmethod
    def _record_to_response(record: DatasetRecord) -> DatasetResponse:
        """Convert a DatasetRecord to a DatasetResponse."""
        schema_columns = [SchemaColumn(**col) for col in json.loads(record.schema_json)]
        return DatasetResponse(
            id=record.id,
            name=record.name,
            description=record.description,
            table_name=record.table_name,
            row_count=record.row_count,
            schema_info=schema_columns,
            file_size_bytes=record.file_size_bytes,
            created_at=record.created_at,
        )
